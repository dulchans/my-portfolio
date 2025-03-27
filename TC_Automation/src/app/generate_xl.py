import openpyxl as xl
import shutil
import platform
from app.settings import dt, BROWSER, NEW_FILE_PATH, FORMAT_FILE_PATH

def run_generate_xl(judge_results):
    '''
    テストの実行結果を入力したExcelファイルを生成し、テストカバレッジを返す関数
    '''
    # 進行状況を出力
    print('Generating the excel file...')
    
    # 各種情報を取得
    date = dt
    browser = BROWSER
    new_file_path = NEW_FILE_PATH
    format_file_path = FORMAT_FILE_PATH
    
    # フォーマットをコピーして、別名のExcelファイルを生成
    shutil.copyfile(format_file_path, new_file_path)

    # 生成したファイルを開いて、作業シートをアクティブにする
    wb = xl.load_workbook(new_file_path)
    ws = wb['TEST_CASE']
    
    # 各種情報を入力
    ws['E4'] = date                           # テストの実行日
    ws['D6'] = platform.platform()            # テストを実行したPCのOS情報
    ws['E6'] = browser                        # テストを実行したブラウザ
    
    # 判定結果を入力するセル範囲を指定
    ws_judge = ws['D15:D20'] 
    # 判定結果を入力
    for i, row in enumerate(ws_judge):
        for j, cell in enumerate(row):
            # 判定結果の数が対象セル以上の場合は、判定結果を入力
            if i < len(judge_results):
                cell.value = judge_results[i]
            else:
                pass

    # テストカバレッジを算出
    pass_count = judge_results.count('PASS')
    test_coverage = f'{pass_count / len(ws_judge) * 100:.0f}%'
    
    # ファイルを保存
    wb.save(new_file_path)
    print('DONE')

    # 算出したテストカバレッジを返す
    return test_coverage
